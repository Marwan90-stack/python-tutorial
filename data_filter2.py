import os
import pandas as pd
import openpyxl as xl
import time

# print(os.getcwd())
home_dir = "/Users/muhamadmarwan/Documents/Tutorial 1"
path = os.path.join(home_dir, "data")


files = os.listdir(path)

# for file in files:
#     if file.endswith(".xlsx"):
#         print(file)

os.chdir(path)
# print(os.getcwd())

contoh = "Nasional_Tabulasi_UTP_BAB_8_tabel_8_31_komoditas_6212.xlsx"

# file = xl.load_workbook(contoh)
# sheets = pd.Series(file.sheetnames)
# sheet_kab = sheets.loc[sheets.str.contains('kab'), ].reset_index(drop = True)
# sheet_kab = sheet_kab[0]
# sheet_kec = sheets.loc[sheets.str.contains('kec'), ].reset_index(drop = True)
# sheet_kec = sheet_kec[0]

# df_kab = pd.read_excel(contoh, sheet_name = sheet_kab)
# df_kab = df_kab.query('kab == 8101')

# df_kec = pd.read_excel(contoh, sheet_name = sheet_kec)
# df_kec = df_kec.query('kab == 8101')

# folder_output = os.path.join(home_dir, "output")
# # print(folder_output)
# file_output = os.path.join(folder_output, contoh)
# if folder_output is None:
#     raise ValueError("Folder output harus diisi")
# else: 
#     with pd.ExcelWriter(file_output) as writer:
#         df_kab.to_excel(writer, sheet_name=sheet_kab, index=False)
#         df_kec.to_excel(writer, sheet_name=sheet_kec, index=False)

def filter_data(contoh, output_folder = None):
    file = xl.load_workbook(contoh)
    sheets = pd.Series(file.sheetnames)
    sheet_kab = sheets.loc[sheets.str.contains('kab'), ].reset_index(drop = True)
    sheet_kab = sheet_kab[0]
    sheet_kec = sheets.loc[sheets.str.contains('kec'), ].reset_index(drop = True)
    sheet_kec = sheet_kec[0]

    df_kab = pd.read_excel(contoh, sheet_name = sheet_kab)
    df_kab = df_kab.query('kab == 8101')

    df_kec = pd.read_excel(contoh, sheet_name = sheet_kec)
    df_kec = df_kec.query('kab == 8101')

    folder = os.path.join(home_dir, output_folder)
    file_output = os.path.join(folder, contoh)

    if folder is None:
        raise ValueError("Folder output harus diisi")
    else:
        with pd.ExcelWriter(file_output) as writer:
            df_kab.to_excel(writer, sheet_name=sheet_kab, index=False)
            df_kec.to_excel(writer, sheet_name=sheet_kec, index=False)

# files = pd.Series(files)
time_start = time.time()
# files.apply(filter_data, output_folder = 'output')
filter_data(contoh, output_folder = 'output')
time_end = time.time()
durasi = time_end - time_start
print(f"waktu yang dibutuhkan untuk filter data selama {durasi} detik")

